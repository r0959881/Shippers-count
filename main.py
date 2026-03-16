import math
import re
import os
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


FILL_THRESHOLD = 0.80


# ==========================
# CALCULATION FUNCTIONS
# ==========================
def calc_single(shipper, product):
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / a)
    nB = math.floor(B / b)
    nC = math.floor(C / c)
    qty = nA * nB * nC
    return nA, nB, nC, qty


def calc_wrap_option1(shipper, product, N):
    # Option 1 = N·A × B × C => (N*a, b, c)
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / (N * a))
    nB = math.floor(B / b)
    nC = math.floor(C / c)
    bundles = nA * nB * nC
    qty = bundles * N
    return nA, nB, nC, bundles, qty


def calc_wrap_option2(shipper, product, N):
    # Option 2 = A × N·B × C => (a, N*b, c)
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / a)
    nB = math.floor(B / (N * b))
    nC = math.floor(C / c)
    bundles = nA * nB * nC
    qty = bundles * N
    return nA, nB, nC, bundles, qty


def fill_percent(qty, shipper, product):
    A, B, C = shipper
    a, b, c = product
    shipper_vol = A * B * C
    product_vol = a * b * c
    return (qty * product_vol) / shipper_vol if shipper_vol > 0 else 0.0


# ==========================
# EXCEL LOADER (robust)
# ==========================
def load_shippers_from_excel(filepath: str) -> pd.DataFrame:
    raw = pd.read_excel(filepath, sheet_name=0, header=None)

    header_row_idx = None
    for i in range(min(120, len(raw))):
        row_vals = raw.iloc[i].astype(str).str.strip().tolist()
        if ("A" in row_vals) and ("B" in row_vals) and ("C" in row_vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row containing A, B, C in the shipper Excel.")

    df = pd.read_excel(filepath, sheet_name=0, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    if "A" not in df.columns or "B" not in df.columns or "C" not in df.columns:
        raise ValueError(f"Missing A/B/C columns. Found: {df.columns.tolist()}")

    candidate_name_cols = [c for c in df.columns if c not in ["A", "B", "C"]]
    name_col = candidate_name_cols[0] if candidate_name_cols else None

    df["A_num"] = pd.to_numeric(df["A"], errors="coerce")
    df["B_num"] = pd.to_numeric(df["B"], errors="coerce")
    df["C_num"] = pd.to_numeric(df["C"], errors="coerce")

    if name_col is None:
        df["Shipper"] = [f"Shipper_{i}" for i in range(len(df))]
        name_col = "Shipper"
    else:
        df[name_col] = df[name_col].astype(str).str.strip()

    clean = df.dropna(subset=["A_num", "B_num", "C_num"]).copy()
    clean = clean[(clean["A_num"] > 0) & (clean["B_num"] > 0) & (clean["C_num"] > 0)]

    def looks_like_shipper(s: str) -> bool:
        s_low = str(s).lower()
        if "shipper" in s_low:
            return True
        return bool(re.match(r"^[A-Za-z0-9][A-Za-z0-9 \-_/]+$", str(s))) and len(str(s)) >= 3

    clean = clean[clean[name_col].apply(looks_like_shipper)]

    out = pd.DataFrame({
        "Shipper": clean[name_col],
        "A": clean["A_num"].astype(int),
        "B": clean["B_num"].astype(int),
        "C": clean["C_num"].astype(int),
    }).reset_index(drop=True)

    if out.empty:
        raise ValueError("No valid shipper rows found after cleaning.")
    return out


# ==========================
# EXCEL FORMATTING
# ==========================
def format_sheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws[f"{col_letter}{row}"].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def apply_good_bad_colors(ws, status_col="Status"):
    header = [c.value for c in ws[1]]
    if status_col not in header:
        return
    idx = header.index(status_col) + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=idx)
        if str(cell.value).upper() == "GOOD":
            cell.fill = green_fill
        else:
            cell.fill = red_fill


# ==========================
# REPORT GENERATION
# ==========================
def generate_report(shipper_excel_path: str, product: tuple[float, float, float], N: int, out_path: str):
    shippers = load_shippers_from_excel(shipper_excel_path)

    use_wrap = N > 1

    single_rows, opt1_rows, opt2_rows, summary_rows = [], [], [], []

    for _, r in shippers.iterrows():
        shipper = (r["A"], r["B"], r["C"])

        # Single
        nA, nB, nC, qty = calc_single(shipper, product)
        fill = fill_percent(qty, shipper, product)
        single_rows.append({
            "Shipper": r["Shipper"], "A": r["A"], "B": r["B"], "C": r["C"],
            "nA": nA, "nB": nB, "nC": nC,
            "Total Qty": qty,
            "Fill %": round(fill * 100, 2),
            "Status": "GOOD" if fill >= FILL_THRESHOLD else "NOT GOOD"
        })

        # Wrap options only when N>1
        if use_wrap:
            nA1, nB1, nC1, bundles1, qty1 = calc_wrap_option1(shipper, product, N)
            fill1 = fill_percent(qty1, shipper, product)
            opt1_rows.append({
                "Shipper": r["Shipper"], "A": r["A"], "B": r["B"], "C": r["C"],
                "nA": nA1, "nB": nB1, "nC": nC1,
                "Bundles": bundles1,
                "Total Qty": qty1,
                "Fill %": round(fill1 * 100, 2),
                "Status": "GOOD" if fill1 >= FILL_THRESHOLD else "NOT GOOD"
            })

            nA2, nB2, nC2, bundles2, qty2 = calc_wrap_option2(shipper, product, N)
            fill2 = fill_percent(qty2, shipper, product)
            opt2_rows.append({
                "Shipper": r["Shipper"], "A": r["A"], "B": r["B"], "C": r["C"],
                "nA": nA2, "nB": nB2, "nC": nC2,
                "Bundles": bundles2,
                "Total Qty": qty2,
                "Fill %": round(fill2 * 100, 2),
                "Status": "GOOD" if fill2 >= FILL_THRESHOLD else "NOT GOOD"
            })

        # Summary (best passing >=80% else best qty)
        if use_wrap:
            candidates = [
                ("Single", qty, fill, (nA, nB, nC)),
                (f"Wrap Option 1 (N*A) N={N}", qty1, fill1, (nA1, nB1, nC1)),
                (f"Wrap Option 2 (N*B) N={N}", qty2, fill2, (nA2, nB2, nC2)),
            ]
        else:
            candidates = [("Single", qty, fill, (nA, nB, nC))]

        passing = [x for x in candidates if x[2] >= FILL_THRESHOLD]
        best = max(passing, key=lambda x: x[1], default=max(candidates, key=lambda x: x[1]))

        summary_rows.append({
            "Shipper": r["Shipper"], "A": r["A"], "B": r["B"], "C": r["C"],
            "Best Mode": best[0],
            "nA": best[3][0], "nB": best[3][1], "nC": best[3][2],
            "Total Qty": best[1],
            "Fill %": round(best[2] * 100, 2),
            "Status": "GOOD" if best[2] >= FILL_THRESHOLD else "NOT GOOD"
        })

    # DataFrames and sorting
    single_df = pd.DataFrame(single_rows).sort_values(["Status", "Fill %", "Total Qty"], ascending=[True, False, False])

    summary_df = pd.DataFrame(summary_rows)
    summary_df["Status_sort"] = summary_df["Status"].apply(lambda x: 0 if x == "GOOD" else 1)
    summary_df = summary_df.sort_values(["Status_sort", "Fill %", "Total Qty"], ascending=[True, False, False]).drop(columns=["Status_sort"])

    if use_wrap:
        opt1_df = pd.DataFrame(opt1_rows).sort_values(["Status", "Fill %", "Total Qty"], ascending=[True, False, False])
        opt2_df = pd.DataFrame(opt2_rows).sort_values(["Status", "Fill %", "Total Qty"], ascending=[True, False, False])

    # Write report
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        single_df.to_excel(writer, sheet_name="Single", index=False)
        if use_wrap:
            opt1_df.to_excel(writer, sheet_name="Wrap_Option_1", index=False)
            opt2_df.to_excel(writer, sheet_name="Wrap_Option_2", index=False)

    # Format
    wb = openpyxl.load_workbook(out_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        format_sheet(ws)
        apply_good_bad_colors(ws, "Status")
    wb.save(out_path)


# ==========================
# GUI
# ==========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Packing Tool")
        self.geometry("520x330")
        self.resizable(False, False)

        self.shipper_path = tk.StringVar(value="")
        self.a_var = tk.StringVar(value="")
        self.b_var = tk.StringVar(value="")
        self.c_var = tk.StringVar(value="")
        self.n_var = tk.StringVar(value="3")

        self._build_ui()

    def _build_ui(self):
        pad = {"padx": 10, "pady": 6}

        tk.Label(self, text="Shipper Excel file (contains A, B, C columns):").pack(anchor="w", **pad)
        row = tk.Frame(self)
        row.pack(fill="x", **pad)

        tk.Entry(row, textvariable=self.shipper_path).pack(side="left", fill="x", expand=True)
        tk.Button(row, text="Browse", command=self.browse_shipper).pack(side="left", padx=6)

        tk.Label(self, text="Product dimensions (mm)").pack(anchor="w", **pad)

        grid = tk.Frame(self)
        grid.pack(fill="x", **pad)

        tk.Label(grid, text="A (Length)").grid(row=0, column=0, sticky="w")
        tk.Entry(grid, textvariable=self.a_var, width=12).grid(row=0, column=1, padx=8)

        tk.Label(grid, text="B (Width)").grid(row=0, column=2, sticky="w")
        tk.Entry(grid, textvariable=self.b_var, width=12).grid(row=0, column=3, padx=8)

        tk.Label(grid, text="C (Height)").grid(row=1, column=0, sticky="w", pady=8)
        tk.Entry(grid, textvariable=self.c_var, width=12).grid(row=1, column=1, padx=8, pady=8)

        tk.Label(grid, text="Wrap quantity N").grid(row=1, column=2, sticky="w", pady=8)
        tk.Entry(grid, textvariable=self.n_var, width=12).grid(row=1, column=3, padx=8, pady=8)

        tk.Label(self, text="Note: If N=1, wrap options are hidden (Single only).").pack(anchor="w", **pad)

        btn_row = tk.Frame(self)
        btn_row.pack(fill="x", **pad)

        tk.Button(btn_row, text="Generate Report", command=self.run).pack(side="left")
        tk.Button(btn_row, text="Exit", command=self.destroy).pack(side="right")

    def browse_shipper(self):
        path = filedialog.askopenfilename(
            title="Select shipper Excel file",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if path:
            self.shipper_path.set(path)

    def run(self):
        try:
            shipper_excel = self.shipper_path.get().strip()
            if not shipper_excel:
                messagebox.showerror("Missing file", "Please select the shipper Excel file.")
                return
            if not os.path.exists(shipper_excel):
                messagebox.showerror("File not found", "Selected shipper Excel file does not exist.")
                return

            a = float(self.a_var.get().strip())
            b = float(self.b_var.get().strip())
            c = float(self.c_var.get().strip())
            N = int(self.n_var.get().strip())
            if N < 1:
                N = 1

            # Choose where to save
            default_name = f"Packing_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            out_path = filedialog.asksaveasfilename(
                title="Save report as",
                defaultextension=".xlsx",
                initialfile=default_name,
                filetypes=[("Excel files", "*.xlsx")]
            )
            if not out_path:
                return  # user cancelled

            generate_report(shipper_excel, (a, b, c), N, out_path)

            messagebox.showinfo(
                "Done",
                f"Report created:\n{out_path}\n\nGOOD if Fill% ≥ {int(FILL_THRESHOLD*100)}%"
            )

        except PermissionError as e:
            messagebox.showerror("Permission error", f"Close Excel files and try again.\n\nDetails: {e}")
        except ValueError as e:
            messagebox.showerror("Input error", str(e))
        except Exception as e:
            messagebox.showerror("Error", f"Something went wrong:\n\n{e}")


if __name__ == "__main__":
    app = App()
    app.mainloop()