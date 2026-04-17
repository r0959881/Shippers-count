import math
import re
import os
import json
import tempfile
import webbrowser
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


FILL_THRESHOLD = 0.80
VIEWER_NEAR_FILL_MARGIN = 0.05


def get_weight_limit(line_type: str) -> float:
  line = (line_type or "").strip().lower()
  if line == "robot":
    return 14.0
  return 10.0


def get_line_label(line_type: str) -> str:
  return "Robot line (14 kg)" if (line_type or "").strip().lower() == "robot" else "Non-robot line (10 kg)"


def calc_total_weight(qty: int, piece_weight: float) -> float:
  return float(qty) * float(piece_weight)


def calc_weight_utilization(total_weight: float, weight_limit: float) -> float:
  return (float(total_weight) / float(weight_limit)) if weight_limit > 0 else 0.0


def calc_weight_gap(total_weight: float, weight_limit: float) -> float:
  return float(weight_limit) - float(total_weight)


def get_weight_status(total_weight: float, weight_limit: float) -> str:
  return "GOOD" if total_weight <= weight_limit else "EXCEEDED WEIGHT"


def make_weight_metrics(total_weight: float, weight_limit: float) -> dict:
  utilization = calc_weight_utilization(total_weight, weight_limit)
  gap = calc_weight_gap(total_weight, weight_limit)
  return {
    "Weight % of Limit": round(utilization * 100, 2),
    "Remaining to Limit (kg)": round(max(gap, 0.0), 4),
    "Over Limit (kg)": round(max(-gap, 0.0), 4),
  }


def status_rank(status: str) -> int:
  status = str(status).upper()
  if status == "GOOD":
    return 0
  if status == "UNDER LIMIT":
    return 1
  return 2


# ==========================
# CALCULATION FUNCTIONS
# ==========================
def calc_single(shipper, product):
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / a)
    nB = math.floor(B / b)
    nC = math.floor(C / c)
    qty = nA * nB * nC
    return nA, nB, nC, qty


def calc_wrap_option1(shipper, product, N):
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / (N * a))
    nB = math.floor(B / b)
    nC = math.floor(C / c)
    bundles = nA * nB * nC
    qty = bundles * N
    return nA, nB, nC, bundles, qty


def calc_wrap_option2(shipper, product, N):
    A, B, C = shipper
    a, b, c = product
    nA = math.floor(A / (N * b))
    nB = math.floor(B / a)
    nC = math.floor(C / c)
    bundles = nA * nB * nC
    qty = bundles * N
    return nA, nB, nC, bundles, qty


def fill_percent(qty, shipper, product):
    A, B, C = shipper
    a, b, c = product
    shipper_vol = A * B * C
    product_vol = a * b * c
    return (qty * product_vol) / shipper_vol if shipper_vol > 0 else 0.0


# ==========================
# EXCEL LOADER (robust)
# ==========================
def load_shippers_from_excel(filepath: str) -> pd.DataFrame:
    raw = pd.read_excel(filepath, sheet_name=0, header=None)

    header_row_idx = None
    for i in range(min(120, len(raw))):
        row_vals = raw.iloc[i].astype(str).str.strip().tolist()
        if ("A" in row_vals) and ("B" in row_vals) and ("C" in row_vals):
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Could not find header row containing A, B, C in the shipper Excel.")

    df = pd.read_excel(filepath, sheet_name=0, header=header_row_idx)
    df.columns = [str(c).strip() for c in df.columns]

    if "A" not in df.columns or "B" not in df.columns or "C" not in df.columns:
        raise ValueError(f"Missing A/B/C columns. Found: {df.columns.tolist()}")

    candidate_name_cols = [c for c in df.columns if c not in ["A", "B", "C"]]
    name_col = candidate_name_cols[0] if candidate_name_cols else None

    df["A_num"] = pd.to_numeric(df["A"], errors="coerce")
    df["B_num"] = pd.to_numeric(df["B"], errors="coerce")
    df["C_num"] = pd.to_numeric(df["C"], errors="coerce")

    if name_col is None:
        df["Shipper"] = [f"Shipper_{i}" for i in range(len(df))]
        name_col = "Shipper"
    else:
        df[name_col] = df[name_col].astype(str).str.strip()

    clean = df.dropna(subset=["A_num", "B_num", "C_num"]).copy()
    clean = clean[(clean["A_num"] > 0) & (clean["B_num"] > 0) & (clean["C_num"] > 0)]

    def looks_like_shipper(s: str) -> bool:
        s_low = str(s).lower()
        if "shipper" in s_low:
            return True
        return bool(re.match(r"^[A-Za-z0-9][A-Za-z0-9 \-_/]+$", str(s))) and len(str(s)) >= 3

    clean = clean[clean[name_col].apply(looks_like_shipper)]

    out = pd.DataFrame({
        "Shipper": clean[name_col],
        "A": clean["A_num"].astype(int),
        "B": clean["B_num"].astype(int),
        "C": clean["C_num"].astype(int),
    }).reset_index(drop=True)

    if out.empty:
        raise ValueError("No valid shipper rows found after cleaning.")
    return out


# ==========================
# EXCEL FORMATTING
# ==========================
def format_sheet(ws):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            val = ws[f"{col_letter}{row}"].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 55)


def apply_good_bad_colors(ws, status_col="Status"):
    header = [c.value for c in ws[1]]
    if status_col not in header:
        return
    idx = header.index(status_col) + 1

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=idx)
        status = str(cell.value).upper()
        if status == "GOOD":
            cell.fill = green_fill
        elif status == "UNDER LIMIT":
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            cell.fill = red_fill


# ==========================
# 3D HTML VIEWER GENERATOR
# ==========================
def generate_html_viewer(good_results: list, product: tuple, product_name: str = "") -> str:
    """
    good_results: list of dicts with keys:
        shipper_name, sA, sB, sC, pA, pB, pC, nA, nB, nC, qty, fill, mode, N
    product: (a, b, c) in mm
    Returns HTML string.
    """
    pA, pB, pC = product
    data_json = json.dumps(good_results)
    header_label = product_name.strip() if product_name.strip() else f"{pA} × {pB} × {pC} mm"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Packing Results — Good Configurations (≥80% fill)</title>
<style>
  :root {{
    --bg: #0f1117;
    --surface: #1a1d27;
    --surface2: #22263a;
    --border: rgba(255,255,255,0.08);
    --accent: #4f8ef7;
    --accent2: #f7934f;
    --green: #3dd68c;
    --text: #e8eaf2;
    --muted: #7a7f9a;
    --font: 'DM Sans', system-ui, sans-serif;
    --mono: 'DM Mono', monospace;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    background: var(--bg);
    color: var(--text);
    font-family: var(--font);
    min-height: 100vh;
  }}
  @import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600&family=DM+Mono:wght@400;500&display=swap');

  .header {{
    padding: 28px 32px 20px;
    border-bottom: 1px solid var(--border);
    display: flex;
    align-items: baseline;
    gap: 16px;
    flex-wrap: wrap;
  }}
  .header h1 {{
    font-size: 22px;
    font-weight: 600;
    letter-spacing: -0.3px;
  }}
  .header .sub {{
    color: var(--muted);
    font-size: 13px;
    font-family: var(--mono);
  }}
  .product-pill {{
    margin-left: auto;
    background: var(--surface2);
    border: 1px solid var(--border);
    border-radius: 8px;
    padding: 6px 14px;
    font-family: var(--mono);
    font-size: 13px;
    color: var(--accent);
  }}

  .layout {{
    display: flex;
    height: calc(100vh - 73px);
  }}

  /* Sidebar */
  .sidebar {{
    width: 300px;
    min-width: 280px;
    border-right: 1px solid var(--border);
    overflow-y: auto;
    padding: 12px 8px;
    flex-shrink: 0;
  }}
  .sidebar-title {{
    font-size: 11px;
    font-weight: 500;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--muted);
    padding: 4px 10px 10px;
  }}
  .shipper-card {{
    border-radius: 10px;
    padding: 12px 14px;
    cursor: pointer;
    transition: background 0.15s;
    border: 1px solid transparent;
    margin-bottom: 4px;
  }}
  .shipper-card:hover {{
    background: var(--surface);
  }}
  .shipper-card.active {{
    background: var(--surface2);
    border-color: var(--accent);
  }}
  .shipper-card .name {{
    font-size: 14px;
    font-weight: 500;
    margin-bottom: 4px;
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
  }}
  .shipper-card .dims {{
    font-family: var(--mono);
    font-size: 11px;
    color: var(--muted);
    margin-bottom: 6px;
  }}
  .shipper-card .row {{
    display: flex;
    align-items: center;
    justify-content: space-between;
  }}
  .shipper-card .qty {{
    font-size: 13px;
    font-weight: 600;
    color: var(--green);
  }}
  .fill-bar-wrap {{
    flex: 1;
    margin-left: 12px;
    height: 4px;
    background: rgba(255,255,255,0.08);
    border-radius: 4px;
    overflow: hidden;
  }}
  .fill-bar {{
    height: 100%;
    border-radius: 4px;
    background: linear-gradient(90deg, #3dd68c, #4f8ef7);
    transition: width 0.3s;
  }}
  .fill-pct {{
    font-family: var(--mono);
    font-size: 11px;
    color: var(--muted);
    margin-left: 8px;
    min-width: 36px;
    text-align: right;
  }}

  /* Main panel */
  .main {{
    flex: 1;
    display: flex;
    flex-direction: column;
    overflow: hidden;
  }}
  .canvas-wrap {{
    flex: 1;
    display: flex;
    align-items: center;
    justify-content: center;
    position: relative;
    min-height: 0;
  }}
  canvas {{
    display: block;
    cursor: grab;
    max-width: 100%;
    max-height: 100%;
  }}
  canvas:active {{ cursor: grabbing; }}
  .drag-hint {{
    position: absolute;
    bottom: 14px;
    right: 18px;
    font-size: 11px;
    color: var(--muted);
    pointer-events: none;
  }}

  /* Stats bar */
  .stats-bar {{
    border-top: 1px solid var(--border);
    padding: 14px 28px;
    display: flex;
    gap: 32px;
    flex-wrap: wrap;
    background: var(--surface);
  }}
  .stat {{
    display: flex;
    flex-direction: column;
    gap: 2px;
  }}
  .stat .label {{
    font-size: 11px;
    color: var(--muted);
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }}
  .stat .value {{
    font-family: var(--mono);
    font-size: 18px;
    font-weight: 500;
    color: var(--text);
  }}
  .stat .value.green {{ color: var(--green); }}
  .stat .value.accent {{ color: var(--accent); }}
  .stat .value.accent2 {{ color: var(--accent2); }}

  .no-results {{
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 100%;
    color: var(--muted);
    gap: 12px;
    font-size: 15px;
  }}
</style>
</head>
<body>

<div class="header">
  <h1>Good Packing Results</h1>
  <span class="sub" id="good-count"></span>
  <div class="product-pill" id="prod-label">Product: {header_label}</div>
</div>

<div class="layout">
  <div class="sidebar">
    <div class="sidebar-title">Configurations: >=80% fill + near-threshold under-limit</div>
    <div id="list"></div>
  </div>
  <div class="main">
    <div class="canvas-wrap">
      <canvas id="c"></canvas>
      <div class="drag-hint">drag to rotate · scroll to zoom</div>
    </div>
    <div class="stats-bar" id="statsbar"></div>
  </div>
</div>

<script>
const DATA = {data_json};

let current = 0;
let rotX = -0.42, rotY = 0.62, scale = 1;
let dragging = false, lastX, lastY;

const cv = document.getElementById('c');
const ctx = cv.getContext('2d');

function resizeCanvas() {{
  const wrap = cv.parentElement;
  const W = wrap.clientWidth - 32;
  const H = wrap.clientHeight - 32;
  cv.width = Math.max(300, W);
  cv.height = Math.max(200, H);
  draw();
}}
window.addEventListener('resize', resizeCanvas);

cv.addEventListener('mousedown', e => {{ dragging = true; lastX = e.clientX; lastY = e.clientY; }});
document.addEventListener('mouseup', () => dragging = false);
document.addEventListener('mousemove', e => {{
  if (!dragging) return;
  rotY += (e.clientX - lastX) * 0.012;
  rotX += (e.clientY - lastY) * 0.008;
  rotX = Math.max(-Math.PI / 2 + 0.05, Math.min(-0.02, rotX));
  lastX = e.clientX; lastY = e.clientY;
  draw();
}});
cv.addEventListener('wheel', e => {{
  scale *= e.deltaY > 0 ? 0.93 : 1.07;
  scale = Math.max(0.3, Math.min(3.5, scale));
  draw();
  e.preventDefault();
}}, {{ passive: false }});
cv.addEventListener('touchstart', e => {{ dragging = true; lastX = e.touches[0].clientX; lastY = e.touches[0].clientY; }});
document.addEventListener('touchend', () => dragging = false);
document.addEventListener('touchmove', e => {{
  if (!dragging) return;
  rotY += (e.touches[0].clientX - lastX) * 0.012;
  rotX += (e.touches[0].clientY - lastY) * 0.008;
  rotX = Math.max(-Math.PI / 2 + 0.05, Math.min(-0.02, rotX));
  lastX = e.touches[0].clientX; lastY = e.touches[0].clientY;
  draw();
}});

function project(x, y, z, d) {{
  const {{ sA, sB, sC }} = d;
  const baseScale = Math.min(cv.width, cv.height) * 0.68 / Math.max(sA, sB, sC) * scale;
  const dx = x * baseScale - sA / 2 * baseScale;
  const dy = y * baseScale - sB / 2 * baseScale;
  const dz = z * baseScale - sC / 2 * baseScale;
  const cosX = Math.cos(rotX), sinX = Math.sin(rotX);
  const cosY = Math.cos(rotY), sinY = Math.sin(rotY);
  const rx = dx * cosY + dz * sinY;
  const ry_t = -dx * sinY + dz * cosY;
  const ry = dy * cosX - ry_t * sinX;
  const rz = dy * sinX + ry_t * cosX;
  return [cv.width / 2 + rx, cv.height / 2 + ry - rz * 0.18, rz];
}}

function makeFace(pts3d, d, fill, stroke, alpha) {{
  const ps = pts3d.map(p => project(p[0], p[1], p[2], d));
  const avgZ = ps.reduce((s, p) => s + p[2], 0) / ps.length;
  return {{ pts: ps.map(p => [p[0], p[1]]), z: avgZ, fill, stroke, alpha }};
}}

function addBox(x0, y0, z0, dx, dy, dz, d, fc, sc, al) {{
  const x1 = x0+dx, y1 = y0+dy, z1 = z0+dz;
  return [
    makeFace([[x0,y0,z0],[x1,y0,z0],[x1,y1,z0],[x0,y1,z0]], d, fc, sc, al),
    makeFace([[x0,y0,z1],[x1,y0,z1],[x1,y1,z1],[x0,y1,z1]], d, fc, sc, al),
    makeFace([[x0,y0,z0],[x1,y0,z0],[x1,y0,z1],[x0,y0,z1]], d, fc, sc, al),
    makeFace([[x0,y1,z0],[x1,y1,z0],[x1,y1,z1],[x0,y1,z1]], d, fc, sc, al),
    makeFace([[x0,y0,z0],[x0,y1,z0],[x0,y1,z1],[x0,y0,z1]], d, fc, sc, al),
    makeFace([[x1,y0,z0],[x1,y1,z0],[x1,y1,z1],[x1,y0,z1]], d, fc, sc, al),
  ];
}}

function draw() {{
  if (!DATA.length) return;
  const d = DATA[current];
  ctx.clearRect(0, 0, cv.width, cv.height);

  const faces = [];
  const {{ sA, sB, sC, pA, pB, pC, nA, nB, nC, mode, N }} = d;
  const effA = mode === 'wrapA' ? pA * N : mode === 'wrapB' ? pB * N : pA;
  const effB = mode === 'wrapA' ? pB : mode === 'wrapB' ? pA : pB;

  const prodFill   = 'rgba(79,142,247,0.82)';
  const prodStroke = 'rgba(120,175,255,0.95)';
  const wrapFill   = prodFill;
  const wrapStroke = prodStroke;

  const MAX_BOXES = 512;
  const total = nA * nB * nC;
  const skip = total > MAX_BOXES ? Math.ceil(total / MAX_BOXES) : 1;
  let idx = 0;

  for (let ia = 0; ia < nA; ia++) {{
    for (let ib = 0; ib < nB; ib++) {{
      for (let ic = 0; ic < nC; ic++) {{
        idx++;
        if (skip > 1 && idx % skip !== 1) continue;
        const x0 = ia * effA, y0 = ib * effB, z0 = ic * pC;

        if (mode === 'wrapA') {{
          for (let k = 0; k < N; k++)
            faces.push(...addBox(x0 + k * pA, y0, z0, pA, pB, pC, d, wrapFill, wrapStroke, 0.88));
        }} else if (mode === 'wrapB') {{
          for (let k = 0; k < N; k++)
            faces.push(...addBox(x0 + k * pB, y0, z0, pB, pA, pC, d, wrapFill, wrapStroke, 0.88));
        }} else {{
          faces.push(...addBox(x0, y0, z0, pA, pB, pC, d, prodFill, prodStroke, 0.88));
        }}
      }}
    }}
  }}

  // Shipper outline (drawn last = on top after sort, but alpha so products show through)
  faces.push(...addBox(0, 0, 0, sA, sB, sC, d, 'rgba(255,255,255,0.03)', 'rgba(255,255,255,0.35)', 0.5));

  faces.sort((a, b) => a.z - b.z);

  for (const f of faces) {{
    ctx.beginPath();
    ctx.moveTo(f.pts[0][0], f.pts[0][1]);
    for (let i = 1; i < f.pts.length; i++) ctx.lineTo(f.pts[i][0], f.pts[i][1]);
    ctx.closePath();
    ctx.globalAlpha = f.alpha;
    ctx.fillStyle = f.fill;
    ctx.fill();
    ctx.strokeStyle = f.stroke;
    ctx.lineWidth = 0.7;
    ctx.stroke();
    ctx.globalAlpha = 1;
  }}
}}

function updateStats() {{
  if (!DATA.length) return;
  const d = DATA[current];
  const modeLabel = d.mode === 'single' ? 'Single' : d.mode === 'wrapA' ? `Wrap ${{d.N}}×A` : `Wrap ${{d.N}}×B`;
  const totalWeight = Number(d.total_weight || 0).toFixed(3);
  const weightLimit = Number(d.weight_limit || 0).toFixed(1);
  const weightPct = Number(d['Weight % of Limit'] || 0).toFixed(1);
  const weightStatus = (d.weight_status || '').toUpperCase();
  const weightColor = weightStatus === 'EXCEEDED WEIGHT' ? '#ff7d8f' : '#3dd68c';
  document.getElementById('statsbar').innerHTML = `
    <div class="stat">
      <span class="label">Shipper</span>
      <span class="value accent" style="font-size:15px">${{d.shipper_name}}</span>
    </div>
    <div class="stat">
      <span class="label">Box (A × B × C)</span>
      <span class="value accent2">${{d.sA}} × ${{d.sB}} × ${{d.sC}} mm</span>
    </div>
    <div class="stat">
      <span class="label">Product (a × b × c)</span>
      <span class="value">${{d.pA}} × ${{d.pB}} × ${{d.pC}} mm</span>
    </div>
    <div class="stat">
      <span class="label">Grid (nA × nB × nC)</span>
      <span class="value">${{d.nA}} × ${{d.nB}} × ${{d.nC}}</span>
    </div>
    <div class="stat">
      <span class="label">Total pieces</span>
      <span class="value green">${{d.qty}}</span>
    </div>
    <div class="stat">
      <span class="label">Fill</span>
      <span class="value green">${{(d.fill * 100).toFixed(1)}}%</span>
    </div>
    <div class="stat">
      <span class="label">Weight</span>
      <span class="value" style="color:${{weightColor}}">${{totalWeight}} / ${{weightLimit}} kg</span>
    </div>
    <div class="stat">
      <span class="label">Weight Fill</span>
      <span class="value">${{weightPct}}%</span>
    </div>
    <div class="stat">
      <span class="label">Mode</span>
      <span class="value" style="font-size:14px">${{modeLabel}}</span>
    </div>
  `;
}}

function buildSidebar() {{
  const list = document.getElementById('list');
  list.innerHTML = '';
  DATA.forEach((d, i) => {{
    const pct = (d.fill * 100).toFixed(1);
    const barW = Math.min(100, d.fill * 100).toFixed(1);
    const wt = Number(d.total_weight || 0).toFixed(3);
    const lim = Number(d.weight_limit || 0).toFixed(1);
    const wtStatus = (d.weight_status || '').toUpperCase();
    const card = document.createElement('div');
    card.className = 'shipper-card' + (i === 0 ? ' active' : '');
    card.innerHTML = `
      <div class="name">${{d.shipper_name}}</div>
      <div class="dims">Mode: ${{d.mode === 'single' ? 'Single' : d.mode === 'wrapA' ? `Wrap ${{d.N}}×A` : `Wrap ${{d.N}}×B`}}</div>
      <div class="dims">${{d.sA}} × ${{d.sB}} × ${{d.sC}} mm</div>
      <div class="dims">Weight: ${{wt}} / ${{lim}} kg · ${{wtStatus}}</div>
      <div class="row">
        <span class="qty">${{d.qty}} pcs</span>
        <div class="fill-bar-wrap"><div class="fill-bar" style="width:${{barW}}%"></div></div>
        <span class="fill-pct">${{pct}}%</span>
      </div>
    `;
    card.addEventListener('click', () => {{
      document.querySelectorAll('.shipper-card').forEach(c => c.classList.remove('active'));
      card.classList.add('active');
      current = i;
      updateStats();
      draw();
    }});
    list.appendChild(card);
  }});
  document.getElementById('good-count').textContent = DATA.length + ' result' + (DATA.length !== 1 ? 's' : '') + ' found';
}}

if (DATA.length === 0) {{
  document.querySelector('.main').innerHTML = '<div class="no-results"><span style="font-size:32px">📦</span><span>No shippers reach 80% fill with these dimensions.</span></div>';
}} else {{
  buildSidebar();
  updateStats();
  resizeCanvas();
}}
</script>
</body>
</html>"""
    return html


# ==========================
# REPORT + VIEWER COMBINED
# ==========================
def compute_results(
  shippers: pd.DataFrame,
  product: tuple,
  N: int,
  selected_wrap_modes: list[str] | None = None,
  piece_weight: float = 0.0,
  line_type: str = "non-robot",
):
  weight_limit = get_weight_limit(line_type)
  use_wrap = N > 1
  if selected_wrap_modes is None:
    selected_wrap_modes = ["single"] if not use_wrap else ["wrapA", "wrapB"]
  selected_wrap_modes = set(selected_wrap_modes)

  good_results = []
  best_results = []
  all_rows = []

  for _, r in shippers.iterrows():
    shipper = (int(r["A"]), int(r["B"]), int(r["C"]))
    pA, pB, pC = product

    shipper_candidates = []

    nA, nB, nC, qty = calc_single(shipper, product)
    fill = fill_percent(qty, shipper, product)
    total_weight = calc_total_weight(qty, piece_weight)
    weight_metrics = make_weight_metrics(total_weight, weight_limit)
    row = dict(
      shipper_name=r["Shipper"], sA=int(r["A"]), sB=int(r["B"]), sC=int(r["C"]),
      pA=pA, pB=pB, pC=pC, nA=nA, nB=nB, nC=nC, qty=qty,
      fill=round(fill, 4), mode="single", N=1,
      piece_weight=round(piece_weight, 4), total_weight=round(total_weight, 4),
      weight_limit=round(weight_limit, 4), line_type=line_type,
      weight_status=get_weight_status(total_weight, weight_limit),
      **weight_metrics,
    )
    all_rows.append(row)
    shipper_candidates.append(row)

    if use_wrap and "wrapA" in selected_wrap_modes:
      nA1, nB1, nC1, bundles1, qty1 = calc_wrap_option1(shipper, product, N)
      fill1 = fill_percent(qty1, shipper, product)
      total_weight1 = calc_total_weight(qty1, piece_weight)
      weight_metrics1 = make_weight_metrics(total_weight1, weight_limit)
      row1 = dict(
        shipper_name=r["Shipper"], sA=int(r["A"]), sB=int(r["B"]), sC=int(r["C"]),
        pA=pA, pB=pB, pC=pC, nA=nA1, nB=nB1, nC=nC1, qty=qty1,
        fill=round(fill1, 4), mode="wrapA", N=N,
        piece_weight=round(piece_weight, 4), total_weight=round(total_weight1, 4),
        weight_limit=round(weight_limit, 4), line_type=line_type,
        weight_status=get_weight_status(total_weight1, weight_limit),
        **weight_metrics1,
      )
      all_rows.append(row1)
      shipper_candidates.append(row1)

    if use_wrap and "wrapB" in selected_wrap_modes:
      nA2, nB2, nC2, bundles2, qty2 = calc_wrap_option2(shipper, product, N)
      fill2 = fill_percent(qty2, shipper, product)
      total_weight2 = calc_total_weight(qty2, piece_weight)
      weight_metrics2 = make_weight_metrics(total_weight2, weight_limit)
      row2 = dict(
        shipper_name=r["Shipper"], sA=int(r["A"]), sB=int(r["B"]), sC=int(r["C"]),
        pA=pA, pB=pB, pC=pC, nA=nA2, nB=nB2, nC=nC2, qty=qty2,
        fill=round(fill2, 4), mode="wrapB", N=N,
        piece_weight=round(piece_weight, 4), total_weight=round(total_weight2, 4),
        weight_limit=round(weight_limit, 4), line_type=line_type,
        weight_status=get_weight_status(total_weight2, weight_limit),
        **weight_metrics2,
      )
      all_rows.append(row2)
      shipper_candidates.append(row2)

    valid_candidates = [
      c for c in shipper_candidates
      if c["qty"] > 0 and c["nA"] > 0 and c["nB"] > 0 and c["nC"] > 0
    ]
    if not valid_candidates:
      continue

    under_limit_candidates = [c for c in valid_candidates if c["total_weight"] <= weight_limit]
    if under_limit_candidates:
      best = max(under_limit_candidates, key=lambda x: (x["fill"], x["qty"], x["Weight % of Limit"]))
    else:
      best = min(valid_candidates, key=lambda x: (x["total_weight"] - weight_limit, -x["fill"], -x["qty"]))
    best_results.append(best)

  good_results = [
    r for r in all_rows
    if r["qty"] > 0 and r["nA"] > 0 and r["nB"] > 0 and r["nC"] > 0
    and r["fill"] >= FILL_THRESHOLD and r["total_weight"] <= weight_limit
  ]

  good_results.sort(key=lambda x: (-x["fill"], -x["qty"]))
  best_results.sort(key=lambda x: (x["total_weight"] > weight_limit, abs(x["total_weight"] - weight_limit), -x["fill"], -x["qty"]))
  all_rows.sort(key=lambda x: (x["total_weight"] > weight_limit, abs(x["total_weight"] - weight_limit), -x["fill"], -x["qty"]))
  return good_results, best_results, all_rows


def get_viewer_results(good_results: list, best_results: list, all_results: list):
    """
    Viewer should show all >= threshold under-limit configurations.
    It also includes under-limit configurations near the fill threshold.
    If still empty, show best-per-shipper fallback.
    If still empty, fall back to the single best overall result.
    """
    near_fill_threshold = max(0.0, FILL_THRESHOLD - VIEWER_NEAR_FILL_MARGIN)
    near_under_limit = [
      r for r in all_results
      if r["qty"] > 0 and r["nA"] > 0 and r["nB"] > 0 and r["nC"] > 0
      and r["total_weight"] <= r["weight_limit"]
      and near_fill_threshold <= r["fill"] < FILL_THRESHOLD
    ]
    near_under_limit.sort(key=lambda x: (abs(FILL_THRESHOLD - x["fill"]), -x["fill"], -x["qty"]))

    if good_results:
      merged = list(good_results)
      existing = {
        (r["shipper_name"], r["mode"], r["N"], r["qty"], r["fill"], r["total_weight"])
        for r in merged
      }
      for r in near_under_limit:
        key = (r["shipper_name"], r["mode"], r["N"], r["qty"], r["fill"], r["total_weight"])
        if key not in existing:
          merged.append(r)
      return merged, False

    if near_under_limit:
      return near_under_limit, True
    if best_results:
      return best_results, True
    if all_results:
        return [all_results[0]], True
    return [], True


def open_3d_viewer(good_results: list, product: tuple, product_name: str = ""):
    html = generate_html_viewer(good_results, product, product_name)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".html", mode="w", encoding="utf-8")
    tmp.write(html)
    tmp.close()
    webbrowser.open(f"file://{tmp.name}")


def generate_report(shipper_excel_path: str, product: tuple, N: int, out_path: str, selected_wrap_modes: list[str] | None = None, product_name: str = "", piece_weight: float = 0.0, line_type: str = "non-robot"):
  shippers = load_shippers_from_excel(shipper_excel_path)
  use_wrap = N > 1
  if selected_wrap_modes is None:
    selected_wrap_modes = ["single"] if not use_wrap else ["wrapA", "wrapB"]
  selected_wrap_modes = set(selected_wrap_modes)
  weight_limit = get_weight_limit(line_type)
  line_label = get_line_label(line_type)

  single_rows, opt1_rows, opt2_rows, summary_rows = [], [], [], []

  for _, r in shippers.iterrows():
    shipper = (r["A"], r["B"], r["C"])

    nA, nB, nC, qty = calc_single(shipper, product)
    fill = fill_percent(qty, shipper, product)
    total_weight = calc_total_weight(qty, piece_weight)
    weight_metrics = make_weight_metrics(total_weight, weight_limit)
    single_rows.append({
      "Shipper": r["Shipper"], "A (mm)": r["A"], "B (mm)": r["B"], "C (mm)": r["C"],
      "nA": nA, "nB": nB, "nC": nC,
      "Total Qty": qty,
      "Fill %": round(fill * 100, 2),
      "Piece Weight (kg)": round(piece_weight, 4),
      "Total Weight (kg)": round(total_weight, 4),
      "Weight Limit (kg)": weight_limit,
      "Weight % of Limit": weight_metrics["Weight % of Limit"],
      "Remaining to Limit (kg)": weight_metrics["Remaining to Limit (kg)"],
      "Over Limit (kg)": weight_metrics["Over Limit (kg)"],
      "Line Type": line_label,
      "Status": "GOOD" if total_weight <= weight_limit and fill >= FILL_THRESHOLD else ("UNDER LIMIT" if total_weight <= weight_limit else "EXCEEDED WEIGHT")
    })

    wrap1_data = None
    wrap2_data = None

    if use_wrap and "wrapA" in selected_wrap_modes:
      nA1, nB1, nC1, bundles1, qty1 = calc_wrap_option1(shipper, product, N)
      fill1 = fill_percent(qty1, shipper, product)
      total_weight1 = calc_total_weight(qty1, piece_weight)
      weight_metrics1 = make_weight_metrics(total_weight1, weight_limit)
      wrap1_data = (nA1, nB1, nC1, qty1, fill1)
      opt1_rows.append({
        "Shipper": r["Shipper"], "A (mm)": r["A"], "B (mm)": r["B"], "C (mm)": r["C"],
        "nA": nA1, "nB": nB1, "nC": nC1,
        "Bundles": bundles1, "Total Qty": qty1,
        "Fill %": round(fill1 * 100, 2),
        "Piece Weight (kg)": round(piece_weight, 4),
        "Total Weight (kg)": round(total_weight1, 4),
        "Weight Limit (kg)": weight_limit,
        "Weight % of Limit": weight_metrics1["Weight % of Limit"],
        "Remaining to Limit (kg)": weight_metrics1["Remaining to Limit (kg)"],
        "Over Limit (kg)": weight_metrics1["Over Limit (kg)"],
        "Line Type": line_label,
        "Status": "GOOD" if total_weight1 <= weight_limit and fill1 >= FILL_THRESHOLD else ("UNDER LIMIT" if total_weight1 <= weight_limit else "EXCEEDED WEIGHT")
      })

    if use_wrap and "wrapB" in selected_wrap_modes:
      nA2, nB2, nC2, bundles2, qty2 = calc_wrap_option2(shipper, product, N)
      fill2 = fill_percent(qty2, shipper, product)
      total_weight2 = calc_total_weight(qty2, piece_weight)
      weight_metrics2 = make_weight_metrics(total_weight2, weight_limit)
      wrap2_data = (nA2, nB2, nC2, qty2, fill2)
      opt2_rows.append({
        "Shipper": r["Shipper"], "A (mm)": r["A"], "B (mm)": r["B"], "C (mm)": r["C"],
        "nA": nA2, "nB": nB2, "nC": nC2,
        "Bundles": bundles2, "Total Qty": qty2,
        "Fill %": round(fill2 * 100, 2),
        "Piece Weight (kg)": round(piece_weight, 4),
        "Total Weight (kg)": round(total_weight2, 4),
        "Weight Limit (kg)": weight_limit,
        "Weight % of Limit": weight_metrics2["Weight % of Limit"],
        "Remaining to Limit (kg)": weight_metrics2["Remaining to Limit (kg)"],
        "Over Limit (kg)": weight_metrics2["Over Limit (kg)"],
        "Line Type": line_label,
        "Status": "GOOD" if total_weight2 <= weight_limit and fill2 >= FILL_THRESHOLD else ("UNDER LIMIT" if total_weight2 <= weight_limit else "EXCEEDED WEIGHT")
      })

    candidates = [("Single", qty, fill, (nA, nB, nC))]
    if wrap1_data is not None:
      candidates.append((f"Wrap Option 1 (N*A) N={N}", wrap1_data[3], wrap1_data[4], (wrap1_data[0], wrap1_data[1], wrap1_data[2])))
    if wrap2_data is not None:
      candidates.append((f"Wrap Option 2 (N*B) N={N}", wrap2_data[3], wrap2_data[4], (wrap2_data[0], wrap2_data[1], wrap2_data[2])))

    valid_candidates = [x for x in candidates if x[1] > 0 and all(v > 0 for v in x[3])]
    if not valid_candidates:
      valid_candidates = candidates
    best = max(valid_candidates, key=lambda x: (x[2], x[1]))

    summary_rows.append({
      "Shipper": r["Shipper"], "A (mm)": r["A"], "B (mm)": r["B"], "C (mm)": r["C"],
      "Best Mode": best[0],
      "nA": best[3][0], "nB": best[3][1], "nC": best[3][2],
      "Total Qty": best[1],
      "Piece Weight (kg)": round(piece_weight, 4),
      "Total Weight (kg)": round(calc_total_weight(best[1], piece_weight), 4),
      "Weight Limit (kg)": weight_limit,
      "Fill %": round(best[2] * 100, 2),
      "Weight % of Limit": round(calc_weight_utilization(calc_total_weight(best[1], piece_weight), weight_limit) * 100, 2),
      "Remaining to Limit (kg)": round(max(calc_weight_gap(calc_total_weight(best[1], piece_weight), weight_limit), 0.0), 4),
      "Over Limit (kg)": round(max(-calc_weight_gap(calc_total_weight(best[1], piece_weight), weight_limit), 0.0), 4),
      "Line Type": line_label,
      "Status": "GOOD" if calc_total_weight(best[1], piece_weight) <= weight_limit and best[2] >= FILL_THRESHOLD else ("UNDER LIMIT" if calc_total_weight(best[1], piece_weight) <= weight_limit else "EXCEEDED WEIGHT")
    })

  single_df = pd.DataFrame(single_rows)
  single_df["_s"] = single_df["Status"].apply(status_rank)
  single_df = single_df.sort_values(["_s", "Fill %", "Total Qty"], ascending=[True, False, False]).drop(columns=["_s"])
  summary_df = pd.DataFrame(summary_rows)
  summary_df["_s"] = summary_df["Status"].apply(lambda x: 0 if x == "GOOD" else 1)
  summary_df = summary_df.sort_values(["_s", "Fill %", "Total Qty"], ascending=[True, False, False]).drop(columns=["_s"])
  _, _, all_rows = compute_results(shippers, product, N, selected_wrap_modes, piece_weight, line_type)
  weight_analysis_df = pd.DataFrame(all_rows)
  if not weight_analysis_df.empty:
    weight_analysis_df = weight_analysis_df.rename(columns={
      "shipper_name": "Shipper",
      "sA": "A (mm)",
      "sB": "B (mm)",
      "sC": "C (mm)",
      "qty": "Total Qty",
      "fill": "Fill Ratio",
      "mode": "Mode",
      "N": "Wrap Qty N",
      "piece_weight": "Piece Weight (kg)",
      "total_weight": "Total Weight (kg)",
      "weight_limit": "Weight Limit (kg)",
      "weight_status": "Weight Status",
    })
    weight_analysis_df["Fill %"] = (weight_analysis_df["Fill Ratio"] * 100).round(2)
    weight_analysis_df["_s"] = weight_analysis_df["Weight Status"].apply(status_rank)
    weight_analysis_df = weight_analysis_df.sort_values(
      ["_s", "Fill %", "Weight % of Limit", "Total Qty"],
      ascending=[True, False, False, False],
    ).drop(columns=["_s"])

  with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="Summary", index=False)
    single_df.to_excel(writer, sheet_name="Single", index=False)
    if opt1_rows:
      opt1_df = pd.DataFrame(opt1_rows)
      opt1_df["_s"] = opt1_df["Status"].apply(status_rank)
      opt1_df.sort_values(["_s", "Fill %", "Total Qty"], ascending=[True, False, False]).drop(columns=["_s"]).to_excel(writer, sheet_name="Wrap_Option_1", index=False)
    if opt2_rows:
      opt2_df = pd.DataFrame(opt2_rows)
      opt2_df["_s"] = opt2_df["Status"].apply(status_rank)
      opt2_df.sort_values(["_s", "Fill %", "Total Qty"], ascending=[True, False, False]).drop(columns=["_s"]).to_excel(writer, sheet_name="Wrap_Option_2", index=False)
    if not weight_analysis_df.empty:
      weight_analysis_df.to_excel(writer, sheet_name="Weight_Analysis", index=False)

  wb = openpyxl.load_workbook(out_path)
  for sheet in wb.sheetnames:
    ws = wb[sheet]
    format_sheet(ws)
    apply_good_bad_colors(ws, "Weight Status" if sheet == "Weight_Analysis" else "Status")
  wb.save(out_path)


# ==========================
# GUI
# ==========================
class App(tk.Tk):
  def __init__(self):
    super().__init__()
    self.title("Packing Tool")
    self.geometry("620x470")
    self.resizable(False, False)

    self.shipper_path = tk.StringVar(value="")
    self.product_name_var = tk.StringVar(value="")
    self.line_type_var = tk.StringVar(value="non-robot")
    self.piece_weight_var = tk.StringVar(value="")
    self.a_var = tk.StringVar(value="")
    self.b_var = tk.StringVar(value="")
    self.c_var = tk.StringVar(value="")
    self.n_var = tk.StringVar(value="3")
    self.wrap1_var = tk.BooleanVar(value=True)
    self.wrap2_var = tk.BooleanVar(value=True)

    self._shippers_df = None
    self._good_results = None
    self._product = None

    self._build_ui()

  def _build_ui(self):
    pad = {"padx": 12, "pady": 6}

    tk.Label(self, text="Shipper Excel file (contains A, B, C columns):").pack(anchor="w", **pad)
    row = tk.Frame(self)
    row.pack(fill="x", **pad)
    tk.Entry(row, textvariable=self.shipper_path).pack(side="left", fill="x", expand=True)
    tk.Button(row, text="Browse…", command=self.browse_shipper).pack(side="left", padx=6)
    tk.Button(row, text="Load", command=self.load_shippers, bg="#d0eaff").pack(side="left")

    tk.Label(self, text="Product name:").pack(anchor="w", **pad)
    tk.Entry(self, textvariable=self.product_name_var).pack(fill="x", padx=12, pady=(0, 6))

    line_row = tk.Frame(self)
    line_row.pack(fill="x", padx=12, pady=(0, 6))
    tk.Label(line_row, text="Line type:").pack(side="left")
    tk.Radiobutton(line_row, text="Non-robot line (10 kg)", variable=self.line_type_var, value="non-robot").pack(side="left", padx=8)
    tk.Radiobutton(line_row, text="Robot line (14 kg)", variable=self.line_type_var, value="robot").pack(side="left", padx=8)
    tk.Label(self, text="The selected line type controls the weight limit used in the report and viewer.", fg="gray", font=("", 9)).pack(anchor="w", padx=12, pady=(0, 6))

    tk.Label(self, text="Piece weight (kg):").pack(anchor="w", **pad)
    tk.Entry(self, textvariable=self.piece_weight_var).pack(fill="x", padx=12, pady=(0, 6))

    self.loaded_label = tk.Label(self, text="No shippers loaded.", fg="gray", font=("", 9))
    self.loaded_label.pack(anchor="w", padx=12)

    tk.Frame(self, height=1, bg="#ddd").pack(fill="x", padx=12, pady=4)

    tk.Label(self, text="Product dimensions (mm)").pack(anchor="w", **pad)

    grid = tk.Frame(self)
    grid.pack(fill="x", **pad)

    tk.Label(grid, text="a — Length").grid(row=0, column=0, sticky="w")
    tk.Entry(grid, textvariable=self.a_var, width=10).grid(row=0, column=1, padx=8)

    tk.Label(grid, text="b — Width").grid(row=0, column=2, sticky="w")
    tk.Entry(grid, textvariable=self.b_var, width=10).grid(row=0, column=3, padx=8)

    tk.Label(grid, text="c — Height").grid(row=1, column=0, sticky="w", pady=8)
    tk.Entry(grid, textvariable=self.c_var, width=10).grid(row=1, column=1, padx=8, pady=8)

    tk.Label(grid, text="Wrap qty N").grid(row=1, column=2, sticky="w", pady=8)
    tk.Entry(grid, textvariable=self.n_var, width=10).grid(row=1, column=3, padx=8, pady=8)

    tk.Label(self, text="N=1 → single only. N>1 → also checks wrap options.", fg="gray", font=("", 9)).pack(anchor="w", padx=12)

    wrap_row = tk.Frame(self)
    wrap_row.pack(fill="x", padx=12, pady=(0, 6))
    tk.Checkbutton(wrap_row, text="Wrap Option 1", variable=self.wrap1_var).pack(side="left")
    tk.Checkbutton(wrap_row, text="Wrap Option 2", variable=self.wrap2_var).pack(side="left", padx=12)

    tk.Frame(self, height=1, bg="#ddd").pack(fill="x", padx=12, pady=6)

    btn_row = tk.Frame(self)
    btn_row.pack(fill="x", padx=12, pady=4)

    tk.Button(btn_row, text="🗂  Generate Excel Report", command=self.run_report, width=22).pack(side="left", padx=(0, 8))
    self.view3d_btn = tk.Button(
      btn_row,
      text="📦  View 3D Good Shippers",
      command=self.open_viewer,
      width=22,
      bg="#c8f0d8",
      state="disabled",
    )
    self.view3d_btn.pack(side="left")
    tk.Button(btn_row, text="Exit", command=self.destroy).pack(side="right")

    self.result_label = tk.Label(self, text="", fg="darkgreen", font=("", 9, "bold"))
    self.result_label.pack(anchor="w", padx=12, pady=4)

  def browse_shipper(self):
    path = filedialog.askopenfilename(
      title="Select shipper Excel file",
      filetypes=[("Excel files", "*.xlsx *.xls")],
    )
    if path:
      self.shipper_path.set(path)
      self._shippers_df = None
      self.loaded_label.config(text="Not loaded yet — click Load.", fg="gray")
      self.view3d_btn.config(state="disabled")

  def load_shippers(self):
    path = self.shipper_path.get().strip()
    if not path or not os.path.exists(path):
      messagebox.showerror("File not found", "Please select a valid Excel file first.")
      return
    try:
      self._shippers_df = load_shippers_from_excel(path)
      n = len(self._shippers_df)
      self.loaded_label.config(
        text=f"✓  {n} shipper{'s' if n != 1 else ''} loaded from file.",
        fg="darkgreen",
      )
    except Exception as e:
      messagebox.showerror("Load error", str(e))

  def _parse_inputs(self):
    a = float(self.a_var.get().strip())
    b = float(self.b_var.get().strip())
    c = float(self.c_var.get().strip())
    N = self._parse_wrap_qty(self.n_var.get().strip())
    piece_weight = float(self.piece_weight_var.get().strip())
    if piece_weight <= 0:
      raise ValueError("Piece weight must be greater than 0.")
    return (a, b, c), N, piece_weight, self.line_type_var.get().strip()

  def _selected_wrap_modes(self):
    if self.n_var.get().strip() == "1":
      return ["single"]

    modes = []
    if self.wrap1_var.get():
      modes.append("wrapA")
    if self.wrap2_var.get():
      modes.append("wrapB")
    if not modes:
      raise ValueError("Select Wrap Option 1, Wrap Option 2, or both.")
    return modes

  def _parse_wrap_qty(self, raw: str) -> int:
    s = raw.strip().lower().replace(" ", "")
    if not s:
      raise ValueError("Wrap qty N is required.")

    if re.fullmatch(r"\d+", s):
      return max(1, int(s))

    if re.fullmatch(r"\d+([x\*]\d+)+", s):
      parts = [int(p) for p in re.split(r"[x\*]", s)]
      total = 1
      for p in parts:
        total *= p
      return max(1, total)

    raise ValueError("Invalid Wrap qty N. Use values like 3 or 2x3.")

  def _compute(self):
    if self._shippers_df is None:
      messagebox.showerror("No shippers", "Load the shipper Excel file first.")
      return None, None
    product, N, piece_weight, line_type = self._parse_inputs()
    selected_wrap_modes = self._selected_wrap_modes()
    good, _, _ = compute_results(self._shippers_df, product, N, selected_wrap_modes, piece_weight, line_type)
    return good, product

  def run_report(self):
    try:
      if self._shippers_df is None:
        messagebox.showerror("No shippers", "Load the shipper Excel file first.")
        return
      product, N, piece_weight, line_type = self._parse_inputs()
      selected_wrap_modes = self._selected_wrap_modes()

      default_name = f"Packing_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
      out_path = filedialog.asksaveasfilename(
        title="Save report as",
        defaultextension=".xlsx",
        initialfile=default_name,
        filetypes=[("Excel files", "*.xlsx")],
      )
      if not out_path:
        return

      generate_report(
        self.shipper_path.get().strip(),
        product,
        N,
        out_path,
        selected_wrap_modes,
        self.product_name_var.get().strip(),
        piece_weight,
        line_type,
      )

      good, best_results, all_results = compute_results(
        self._shippers_df,
        product,
        N,
        selected_wrap_modes,
        piece_weight,
        line_type,
      )
      viewer_results, used_fallback = get_viewer_results(good, best_results, all_results)
      self._good_results = viewer_results
      self._product = product

      n_good = len(good)
      weight_limit = get_weight_limit(line_type)
      line_label = get_line_label(line_type)
      self.result_label.config(
        text=f"✓  Report saved.  {n_good} configuration{'s' if n_good != 1 else ''} meet fill and weight limit ({line_label})."
      )
      self.view3d_btn.config(state="normal")

      fallback_note = ""
      if used_fallback:
        fallback_note = (
          f"\n\nNo shipper met both fill and weight rules, so the 3D viewer will show the best available result under the {weight_limit:.0f} kg limit if possible."
        )

      messagebox.showinfo(
        "Done",
        f"Report created:\n{out_path}\n\n"
        f"{n_good} configuration(s) pass the fill and weight rules for {line_label}.\n"
        f"Click '📦 View 3D Good Shippers' to see them visually."
        f"{fallback_note}",
      )
    except PermissionError as e:
      messagebox.showerror("Permission error", f"Close Excel files and try again.\n\n{e}")
    except ValueError as e:
      messagebox.showerror("Input error", str(e))
    except Exception as e:
      messagebox.showerror("Error", f"Something went wrong:\n\n{e}")

  def open_viewer(self):
    try:
      product, N, piece_weight, line_type = self._parse_inputs()
      if self._shippers_df is None:
        messagebox.showerror("No shippers", "Load the shipper Excel file first.")
        return
      good, best_results, all_results = compute_results(
        self._shippers_df,
        product,
        N,
        self._selected_wrap_modes(),
        piece_weight,
        line_type,
      )
      viewer_results, _ = get_viewer_results(good, best_results, all_results)
      self._good_results = viewer_results
      self._product = product

      if not self._good_results:
        messagebox.showinfo(
          "No results",
          "No valid packing result found with the current dimensions.\n"
          "Try adjusting the product size or wrap quantity N.",
        )
        return

      open_3d_viewer(self._good_results, self._product, self.product_name_var.get().strip())

    except ValueError as e:
      messagebox.showerror("Input error", str(e))
    except Exception as e:
      messagebox.showerror("Error", str(e))


if __name__ == "__main__":
    app = App()
    app.mainloop()
